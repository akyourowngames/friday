"""Resumable research, structured page extraction, and document comparison."""

from __future__ import annotations

import difflib
import html
import hashlib
import json
import re
import sqlite3
import threading
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urljoin, urlparse
from uuid import uuid4

from ares.tools.research import ResearchWorkspace, validate_public_remote_url
from ares.tools.web import fetch_url


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _sentence(value: str, limit: int = 500) -> str:
    cleaned = re.sub(r"\s+", " ", str(value or "")).strip()
    match = re.search(r".+?[.!?](?:\s|$)", cleaned)
    return (match.group(0).strip() if match else cleaned)[:limit]


def _tokens(value: str) -> set[str]:
    return {item.casefold() for item in re.findall(r"[A-Za-z0-9_]{3,}", value)}


def _subqueries(query: str, mode: str, maximum: int, follow_up: str = "") -> list[str]:
    base = f"{query} {follow_up}".strip()
    candidates = [base]
    if mode in {"deep", "fact-check", "compare", "primary", "recommend"}:
        candidates.extend([
            f"{base} primary sources official documentation",
            f"{base} evidence limitations counterarguments",
            f"{base} alternatives comparison",
        ])
    elif mode in {"news", "latest"}:
        candidates.append(f"{base} latest update announcement")
    deduplicated = list(dict.fromkeys(item.strip() for item in candidates if item.strip()))
    return deduplicated[:max(1, min(int(maximum), 8))]


class ResearchUpgradeStore:
    """Persist bounded research graphs and URL snapshots in SQLite."""

    def __init__(self, data_dir: str | Path):
        root = Path(data_dir).expanduser().resolve() / "research"
        root.mkdir(parents=True, exist_ok=True)
        self.root = root
        self._lock = threading.RLock()
        self.conn = sqlite3.connect(
            str(root / "research.db"), timeout=30.0, check_same_thread=False,
        )
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout = 30000")
        self.conn.execute("PRAGMA foreign_keys = ON")
        try:
            self.conn.execute("PRAGMA journal_mode = WAL")
        except sqlite3.DatabaseError:
            pass
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS research_sessions (
                research_id TEXT PRIMARY KEY,
                query TEXT NOT NULL,
                mode TEXT NOT NULL,
                state_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS fetch_snapshots (
                snapshot_id TEXT PRIMARY KEY,
                url TEXT NOT NULL,
                content_hash TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_fetch_snapshot_url ON fetch_snapshots(url, created_at DESC);
            """
        )
        self.conn.commit()

    def close(self) -> None:
        with self._lock:
            self.conn.close()

    def get(self, research_id: str) -> dict[str, Any] | None:
        with self._lock:
            row = self.conn.execute(
                "SELECT state_json FROM research_sessions WHERE research_id = ?", (research_id,),
            ).fetchone()
        if row is None:
            return None
        try:
            value = json.loads(row["state_json"])
        except (TypeError, json.JSONDecodeError):
            return None
        return value if isinstance(value, dict) else None

    def save(self, state: dict[str, Any]) -> None:
        now = _utc_now()
        with self._lock:
            self.conn.execute(
                """INSERT INTO research_sessions (research_id, query, mode, state_json, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?)
                   ON CONFLICT(research_id) DO UPDATE SET
                     query=excluded.query, mode=excluded.mode, state_json=excluded.state_json, updated_at=excluded.updated_at""",
                (
                    state["research_id"], state["query"], state["mode"],
                    json.dumps(state, ensure_ascii=False, sort_keys=True),
                    state.get("created_at") or now, now,
                ),
            )
            self.conn.commit()

    def search(
        self,
        args: dict[str, Any],
        callback: Callable[[dict[str, Any], int], dict[str, Any]],
    ) -> dict[str, Any]:
        mode = str(args.get("search_mode") or "quick").casefold()
        allowed = {"quick", "deep", "news", "fact-check", "compare", "primary", "recommend", "latest"}
        if mode not in allowed:
            raise ValueError(f"search_mode must be one of: {', '.join(sorted(allowed))}")
        research_id = str(args.get("research_id") or uuid4().hex)
        previous = self.get(research_id)
        if args.get("research_id") and previous is None:
            raise ValueError("Unknown research_id")
        query = str(args.get("query") or (previous or {}).get("query") or "").strip()
        if not query:
            raise ValueError("query is required for a new research session")
        follow_up = str(args.get("follow_up") or "").strip()
        queries = _subqueries(query, mode, int(args.get("max_subqueries", 4)), follow_up)
        backend_mode = "news" if mode in {"news", "latest"} else "web"

        def run(item: str) -> tuple[str, dict[str, Any]]:
            request = {
                **args, "query": item, "search_mode": backend_mode,
                "max_results": int(args.get("max_results", 5)),
            }
            request.pop("research_id", None)
            request.pop("follow_up", None)
            try:
                return item, callback(request, int(args.get("fetch_top", 3)))
            except Exception as exc:  # Provider adapters may raise backend-specific errors.
                return item, {
                    "provider": "unknown", "results": [], "source_matrix": [], "fetched": [],
                    "errors": [f"Subquery {item!r} failed: {exc}"],
                }

        with ThreadPoolExecutor(max_workers=min(4, len(queries))) as pool:
            outcomes = list(pool.map(run, queries))

        source_map: dict[str, dict[str, Any]] = {
            str(item["url"]): item for item in (previous or {}).get("sources", []) if item.get("url")
        }
        errors = list((previous or {}).get("errors", []))
        providers: set[str] = set((previous or {}).get("providers", []))
        for subquery, payload in outcomes:
            providers.add(str(payload.get("provider") or "unknown"))
            errors.extend(str(error) for error in payload.get("errors", []))
            matrix = {str(item.get("url")): item for item in payload.get("source_matrix", [])}
            fetched = {str(item.get("url")): item for item in payload.get("fetched", [])}
            for item in payload.get("results", []):
                url = str(item.get("url") or "")
                if not url:
                    continue
                source_id = hashlib.sha256(url.encode()).hexdigest()[:12]
                quality = matrix.get(url, {})
                current = {
                    **item, "source_id": source_id, "subqueries": [subquery],
                    "quality_score": quality.get("quality_score", 0.5),
                    "quality_label": quality.get("quality_label", "standard"),
                    "freshness_label": quality.get("freshness_label", "undated"),
                    "fetched": fetched.get(url, {}),
                }
                if url in source_map:
                    current["subqueries"] = list(dict.fromkeys([*source_map[url].get("subqueries", []), subquery]))
                source_map[url] = current
        minimum_source_score = max(0.0, min(float(args.get("min_source_score", 0.0)), 1.0))
        sources = [
            item for item in source_map.values()
            if float(item.get("quality_score") or 0.0) >= minimum_source_score
        ]
        sources = sorted(sources, key=lambda item: (-float(item.get("rank_score") or item.get("quality_score") or 0), str(item.get("title") or "")))
        if mode == "primary":
            primary = [item for item in sources if item.get("quality_label") in {"authoritative", "primary-or-technical"}]
            sources = primary or sources
        claims: list[dict[str, Any]] = []
        for source in sources:
            claim = _sentence(str(source.get("snippet") or source.get("fetched", {}).get("content") or ""))
            if not claim:
                continue
            overlap_sources = [
                item["source_id"] for item in sources
                if item["source_id"] != source["source_id"] and len(_tokens(claim) & _tokens(str(item.get("snippet") or ""))) >= 3
            ]
            confidence = min(0.98, float(source.get("quality_score") or 0.5) + min(len(overlap_sources), 3) * 0.08)
            candidate = {
                "claim_id": hashlib.sha256((source["source_id"] + claim).encode()).hexdigest()[:12],
                "text": claim, "source_ids": [source["source_id"], *overlap_sources],
                "confidence": round(confidence, 3),
                "inference": False,
            }
            if candidate["confidence"] >= max(0.0, min(float(args.get("min_claim_confidence", 0.0)), 1.0)):
                claims.append(candidate)
        conflicts: list[dict[str, Any]] = []
        for left_index, left in enumerate(claims):
            for right in claims[left_index + 1:]:
                overlap = _tokens(left["text"]) & _tokens(right["text"])
                negated = bool(re.search(r"\b(?:not|no|never|without)\b", left["text"], re.I)) != bool(re.search(r"\b(?:not|no|never|without)\b", right["text"], re.I))
                if len(overlap) >= 4 and negated:
                    conflicts.append({"left_claim_id": left["claim_id"], "right_claim_id": right["claim_id"], "shared_terms": sorted(overlap)[:12]})
        clusters: dict[str, list[str]] = {}
        for source in sources:
            host = str(source.get("host") or urlparse(str(source.get("url") or "")).hostname or "unknown")
            clusters.setdefault(host, []).append(source["source_id"])
        state = {
            "research_id": research_id, "query": query, "mode": mode,
            "created_at": (previous or {}).get("created_at") or _utc_now(), "updated_at": _utc_now(),
            "subqueries": list(dict.fromkeys([*(previous or {}).get("subqueries", []), *queries])),
            "sources": sources, "claims": claims, "clusters": [{"cluster": key, "source_ids": value} for key, value in clusters.items()],
            "conflicts": conflicts, "errors": list(dict.fromkeys(errors)), "providers": sorted(providers),
            "open_questions": ["No independent source corroborates every extracted claim."] if any(len(claim["source_ids"]) == 1 for claim in claims) else [],
            "uncertainty": "Evidence is incomplete." if not claims or errors else "Source-backed claims are available; inspect confidence and conflicts.",
        }
        self.save(state)
        return state

    def snapshot(self, url: str, payload: dict[str, Any]) -> dict[str, Any]:
        content_hash = hashlib.sha256(str(payload.get("content") or "").encode()).hexdigest()
        with self._lock:
            previous_row = self.conn.execute(
                "SELECT snapshot_id, content_hash, payload_json, created_at FROM fetch_snapshots WHERE url = ? ORDER BY created_at DESC LIMIT 1",
                (url,),
            ).fetchone()
            previous = None
            if previous_row:
                try:
                    previous = json.loads(previous_row["payload_json"])
                except (TypeError, json.JSONDecodeError):
                    previous = {}
            snapshot_id = uuid4().hex
            now = _utc_now()
            self.conn.execute(
                "INSERT INTO fetch_snapshots (snapshot_id, url, content_hash, payload_json, created_at) VALUES (?, ?, ?, ?, ?)",
                (snapshot_id, url, content_hash, json.dumps(payload, ensure_ascii=False, sort_keys=True), now),
            )
            self.conn.commit()
            before = str((previous or {}).get("content") or "").splitlines()
            after = str(payload.get("content") or "").splitlines()
            changed = previous_row is not None and previous_row["content_hash"] != content_hash
            summary = list(difflib.unified_diff(before, after, lineterm=""))[:80] if changed else []
            return {
                "snapshot_id": snapshot_id, "content_hash": content_hash, "changed": changed,
                "previous_snapshot_id": previous_row["snapshot_id"] if previous_row else None,
                "previous_created_at": previous_row["created_at"] if previous_row else None,
                "change_summary": summary,
            }


def advanced_fetch(args: dict[str, Any], store: ResearchUpgradeStore) -> dict[str, Any]:
    url = validate_public_remote_url(str(args.get("url") or ""))
    selectors_requested = any(args.get(key) for key in ("selector", "heading", "anchor", "pattern", "extract", "follow_same_domain"))
    page = fetch_url(url, max_chars=int(args.get("max_chars", 15_000)), extract_text=not selectors_requested)
    if page.get("error"):
        return {**page, "selection": {}, "snapshot": store.snapshot(url, page)}
    raw_html = str(page.get("content") or "")
    selection: dict[str, Any] = {}
    if selectors_requested and "html" in str(page.get("content_type") or "").casefold():
        from bs4 import BeautifulSoup

        base_url = str(page.get("final_url") or url)
        soup = BeautifulSoup(raw_html, "html.parser")
        selector = str(args.get("selector") or "").strip()
        heading = str(args.get("heading") or "").strip().casefold()
        anchor = str(args.get("anchor") or "").strip().lstrip("#")
        pattern = str(args.get("pattern") or "")
        nodes = []
        if selector:
            try:
                nodes.extend(soup.select(selector))
            except Exception as exc:
                raise ValueError(f"Invalid CSS selector: {selector}") from exc
        if heading:
            for heading_node in soup.find_all(re.compile(r"^h[1-6]$")):
                if heading not in heading_node.get_text(" ", strip=True).casefold():
                    continue
                nodes.append(heading_node)
                level = int(heading_node.name[1])
                for sibling in heading_node.next_siblings:
                    sibling_name = str(getattr(sibling, "name", "") or "")
                    if re.fullmatch(r"h[1-6]", sibling_name) and int(sibling_name[1]) <= level:
                        break
                    if getattr(sibling, "get_text", None):
                        nodes.append(sibling)
        if anchor:
            anchor_node = soup.find(id=anchor)
            if anchor_node is not None:
                nodes.append(anchor_node)
        nodes = list(dict.fromkeys(nodes))
        selected_text = "\n".join(node.get_text(" ", strip=True) for node in nodes)
        if pattern:
            matches = re.findall(pattern, selected_text or raw_html, re.IGNORECASE | re.MULTILINE)
            selection["pattern_matches"] = matches[:200]
        extract = {str(item).casefold() for item in (args.get("extract") or [])}
        if "links" in extract:
            selection["links"] = [
                {"url": urljoin(base_url, str(item.get("href") or "")), "text": item.get_text(" ", strip=True)}
                for item in soup.find_all("a", href=True)
            ][:500]
        if "meta" in extract:
            selection["meta"] = {
                str(item.get("name") or item.get("property")): str(item.get("content"))
                for item in soup.find_all("meta", content=True)
                if item.get("name") or item.get("property")
            }
        if "json-ld" in extract or "json_ld" in extract:
            json_ld = []
            for item in soup.find_all("script", attrs={"type": re.compile(r"application/ld\+json", re.I)}):
                try:
                    json_ld.append(json.loads(item.get_text()))
                except (TypeError, json.JSONDecodeError):
                    continue
            selection["json_ld"] = json_ld
        if "tables" in extract:
            selection["tables"] = [
                [[cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])] for row in table.find_all("tr")]
                for table in soup.find_all("table")
            ][:50]
        if "headings" in extract:
            selection["headings"] = [
                {"level": int(item.name[1]), "text": item.get_text(" ", strip=True), "id": item.get("id")}
                for item in soup.find_all(re.compile(r"^h[1-6]$"))
            ]
        page["content"] = selected_text[:int(args.get("max_chars", 15_000))] if (selector or heading or anchor) else soup.get_text(" ", strip=True)
        page["content"] = re.sub(r"\s+", " ", page["content"]).strip()
        if bool(args.get("follow_same_domain", False)):
            host = urlparse(url).hostname
            limit = max(0, min(int(args.get("max_follow_pages", 2)), 5))
            followed = []
            seen_links: set[str] = set()
            for link in selection.get("links") or [
                {"url": urljoin(base_url, str(item.get("href") or ""))}
                for item in soup.find_all("a", href=True)
            ]:
                if link["url"] in seen_links or len(followed) >= limit or urlparse(link["url"]).hostname != host:
                    continue
                seen_links.add(link["url"])
                try:
                    validate_public_remote_url(link["url"])
                except ValueError:
                    continue
                child = fetch_url(link["url"], max_chars=min(5_000, int(args.get("max_chars", 15_000))))
                followed.append({"url": link["url"], "title": child.get("title"), "content": child.get("content", ""), "error": child.get("error", "")})
            selection["followed"] = followed
    page["selection"] = selection
    page["language"] = "en" if re.search(r"\b(?:the|and|with|from)\b", str(page.get("content") or ""), re.I) else "unknown"
    page["markdown"] = f"# {page.get('title') or url}\n\n{page.get('content') or ''}"
    if bool(args.get("snapshot", True)) or bool(args.get("compare", False)):
        page["snapshot"] = store.snapshot(url, page)
    return page


def _parse_pages(value: Any, total: int) -> list[int]:
    if not value:
        return list(range(1, total + 1))
    selected: set[int] = set()
    for part in str(value).split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            selected.update(range(max(1, int(start)), min(total, int(end)) + 1))
        elif part:
            selected.add(int(part))
    return sorted(page for page in selected if 1 <= page <= total)


def _extract_one(workspace: ResearchWorkspace, spec: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    path = str(spec.get("path") or "")
    url = str(spec.get("url") or "")
    if url:
        downloaded = workspace.download(url, filename=str(spec.get("filename") or ""), max_bytes=int(args.get("max_bytes", 20 * 1024 * 1024)))
        path = downloaded["path"]
    resolved = Path(path).expanduser().resolve(strict=True)
    suffix = resolved.suffix.casefold()
    if suffix == ".pdf" and args.get("pages"):
        from pypdf import PdfReader
        reader = PdfReader(str(resolved))
        chunks = []
        citations = []
        for page_number in _parse_pages(args.get("pages"), len(reader.pages)):
            text = reader.pages[page_number - 1].extract_text() or ""
            chunks.append(f"## Page {page_number}\n{text}")
            citations.append({"page": page_number, "citation": f"[{resolved.name}#page={page_number}]"})
        result = {"path": str(resolved), "name": resolved.name, "kind": "pdf", "content": "\n\n".join(chunks), "page_citations": citations}
    elif suffix in {".zip", ".jar", ".whl"}:
        with zipfile.ZipFile(resolved) as archive:
            members = archive.infolist()
            total = sum(item.file_size for item in members)
            if len(members) > 10_000 or total > 200 * 1024 * 1024:
                raise ValueError("Archive expansion limits exceeded")
            result = {"path": str(resolved), "name": resolved.name, "kind": "archive", "content": "", "archive_entries": [{"name": item.filename, "bytes": item.file_size, "compressed_bytes": item.compress_size} for item in members[:2_000]]}
    elif suffix == ".xlsx" and (args.get("sheet") or args.get("range")):
        from openpyxl import load_workbook
        workbook = load_workbook(resolved, read_only=True, data_only=True)
        sheet = workbook[str(args.get("sheet") or workbook.sheetnames[0])]
        cells = sheet[str(args.get("range") or sheet.calculate_dimension())]
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        elif cells and not isinstance(cells[0], tuple):
            cells = (cells,)
        rows = [[cell.value for cell in row] for row in cells]
        result = {"path": str(resolved), "name": resolved.name, "kind": "spreadsheet", "content": "\n".join("\t".join("" if value is None else str(value) for value in row) for row in rows), "sheet": sheet.title, "range": str(args.get("range") or sheet.calculate_dimension()), "tables": rows}
    else:
        result = workspace.extract_document(path=str(resolved), max_bytes=int(args.get("max_bytes", 20 * 1024 * 1024)), max_chars=int(args.get("max_chars", 30_000)))
    content = str(result.get("content") or "")
    result["outline"] = [line.strip() for line in content.splitlines() if re.match(r"^#{1,6}\s+", line)]
    result["entities"] = {
        "emails": sorted(set(re.findall(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", content)))[:100],
        "urls": sorted(set(re.findall(r"https?://[^\s)>]+", content)))[:100],
        "dates": sorted(set(re.findall(r"\b\d{4}-\d{2}-\d{2}\b", content)))[:100],
    }
    if bool(args.get("ocr", False)) and not content.strip():
        result["ocr"] = {"attempted": False, "warning": "OCR was requested but no configured OCR provider is available."}
    return result


def advanced_extract(workspace: ResearchWorkspace, args: dict[str, Any]) -> dict[str, Any]:
    specs = list(args.get("documents") or [])
    specs.extend({"path": path} for path in args.get("paths") or [])
    specs.extend({"url": url} for url in args.get("urls") or [])
    if not specs:
        specs = [{"path": args.get("path"), "url": args.get("url"), "filename": args.get("filename")}]
    documents = [_extract_one(workspace, spec, args) for spec in specs]
    mode = str(args.get("mode") or "extract").casefold()
    comparison = []
    if mode == "compare" and len(documents) >= 2:
        left = str(documents[0].get("content") or "").splitlines()
        for document in documents[1:]:
            comparison.append({
                "left": documents[0]["name"], "right": document["name"],
                "diff": list(difflib.unified_diff(left, str(document.get("content") or "").splitlines(), lineterm=""))[:1_000],
            })
    return {"mode": mode, "documents": documents, "comparison": comparison, "document_count": len(documents)}


def create_advanced_report(
    workspace: ResearchWorkspace,
    store: ResearchUpgradeStore,
    args: dict[str, Any],
    callback: Callable[[dict[str, Any], int], dict[str, Any]],
) -> dict[str, Any]:
    style = str(args.get("style") or "brief").casefold()
    allowed = {"brief", "deep", "comparison", "decision", "technical", "market", "study"}
    if style not in allowed:
        raise ValueError(f"style must be one of: {', '.join(sorted(allowed))}")
    research_id = str(args.get("research_id") or "")
    state = store.get(research_id) if research_id else None
    if state is None:
        search_args = {**args, "search_mode": "compare" if style == "comparison" else ("deep" if style in {"deep", "technical", "market", "study", "decision"} else "quick")}
        state = store.search(search_args, callback)
    title = str(args.get("title") or f"Research {style}: {state['query']}").strip()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    workspace.reports_dir.mkdir(parents=True, exist_ok=True)
    output_format = str(args.get("output_format") or "markdown").casefold()
    if output_format not in {"markdown", "json", "html"}:
        raise ValueError("output_format must be markdown, json, or html")
    suffix = {"markdown": ".md", "json": ".json", "html": ".html"}[output_format]
    requested_path = str(args.get("output_path") or "").strip()
    path = (
        Path(requested_path).expanduser().resolve()
        if requested_path
        else workspace.reports_dir / f"{stamp}-{re.sub(r'[^a-z0-9]+', '-', title.casefold()).strip('-')[:80]}{suffix}"
    )
    if path.suffix.casefold() != suffix:
        path = path.with_suffix(suffix)
    path.parent.mkdir(parents=True, exist_ok=True)
    minimum_confidence = max(0.0, min(float(args.get("min_confidence", 0.0)), 1.0))
    maximum_claims = max(1, min(int(args.get("max_claims", 20)), 100))
    report_claims = [
        claim for claim in state.get("claims", [])
        if float(claim.get("confidence") or 0.0) >= minimum_confidence
    ][:maximum_claims]
    sections = {
        "brief": ["Executive summary", "Evidence"],
        "deep": ["Executive summary", "Findings", "Conflicts and limitations"],
        "comparison": ["Comparison frame", "Similarities and differences", "Trade-offs"],
        "decision": ["Decision", "Options", "Recommendation", "Risks"],
        "technical": ["Technical summary", "Architecture and implementation evidence", "Risks"],
        "market": ["Market signal", "Segments and alternatives", "Risks"],
        "study": ["Study notes", "Key concepts", "Open questions"],
    }[style]
    lines = [f"# {title}", "", f"Research ID: `{state['research_id']}`", f"Generated: {_utc_now()}", ""]
    for heading in sections:
        lines.extend([f"## {heading}", ""])
        if heading in {"Executive summary", "Technical summary", "Market signal", "Study notes", "Decision", "Comparison frame"}:
            lines.append(state.get("uncertainty") or "Evidence collected.")
        else:
            for claim in report_claims:
                citations = ", ".join(f"`{source_id}`" for source_id in claim["source_ids"])
                lines.append(f"- {claim['text']} (confidence {claim['confidence']}; sources {citations})")
        lines.append("")
    lines.extend(["## Source map", ""])
    for source in state.get("sources", []):
        lines.append(f"- `{source['source_id']}` [{source.get('title') or source['url']}]({source['url']}) — {source.get('quality_label')}, {source.get('freshness_label')}")
    lines.extend(["", "## Conflicts", ""])
    lines.extend([f"- {json.dumps(item, ensure_ascii=False)}" for item in state.get("conflicts", [])] or ["- No deterministic contradiction was detected."])
    lines.extend(["", "## Open questions", ""])
    lines.extend([f"- {item}" for item in state.get("open_questions", [])] or ["- None recorded."])
    lines.extend(["", "## Methodology", "", f"Mode: {state['mode']}; subqueries: {len(state.get('subqueries', []))}; providers: {', '.join(state.get('providers', []))}."])
    markdown = "\n".join(lines).strip() + "\n"
    if output_format == "json":
        path.write_text(json.dumps({
            "title": title, "style": style, "research": state,
            "included_claims": report_claims, "generated_at": _utc_now(),
        }, ensure_ascii=False, indent=2), encoding="utf-8")
    elif output_format == "html":
        path.write_text(
            "<!doctype html><html><head><meta charset=\"utf-8\"><title>"
            f"{html.escape(title)}</title></head><body><pre>{html.escape(markdown)}</pre></body></html>\n",
            encoding="utf-8",
        )
    else:
        path.write_text(markdown, encoding="utf-8")
    return {
        "path": str(path.resolve()), "name": path.name, "kind": output_format,
        "style": style, "research_id": state["research_id"],
        "sources": len(state.get("sources", [])), "claims": len(report_claims), "state": state,
    }
